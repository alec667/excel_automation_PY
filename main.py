import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]
# dictionary with no duplicates:
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    # add a column: gets the cell object, not the ".value""
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier:
    if supplier_name in products_per_supplier:
        current_num_product = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_product + 1
    else:
        # print("added a new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        currant_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price

inventory_file.save("inventory_with_total_value.xlsx")

print("List of company with respective product count:")
print(products_per_supplier)
print("List of each company with respective total inventory value:")
print(total_value_per_supplier)
print("List of all products with inventory less than 10")
print(products_under_10_inv)
